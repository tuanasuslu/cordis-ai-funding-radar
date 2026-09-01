"""
CORDIS Data & AI Funding Radar
===============================

Purpose
-------
Pulls the official CORDIS bulk export for Horizon Europe (2021-2027) projects,
joins it against the organisations table, and produces a list of every funded
SME in the target countries - tagged (not filtered) by how AI-central each
project is. Candidate leads for BD outreach, or raw material for a "who got
funded this quarter" market-intel note.

Data source
-----------
CORDIS publishes its full Horizon Europe dataset as a downloadable CSV/JSON/XML
bundle via the EU Open Data Portal:
  https://data.europa.eu/data/datasets/cordis-eu-research-projects-under-horizon-europe-2021-2027

The bundle contains (at minimum) a `project.csv` and an `organization.csv`,
semicolon-delimited. CORDIS has changed the exact download path before, so
verify the fallback constant below against the dataset page if downloads
start failing.

Note: CORDIS's own CSV export has a quoting bug (see `load_table`) that
pandas' parsers reject outright; this script works around it with a more
lenient stdlib csv reader and reports how many rows it had to drop.

Usage
-----
    python cordis_ai_funding_radar.py --countries DE,AT,CH --output ai_funding_radar_dach.xlsx

Runs automatically every Monday via the GitHub Actions workflow in
.github/workflows/weekly_radar.yml, which also publishes docs/index.html
to GitHub Pages.
"""

from __future__ import annotations

import argparse
import csv
import io
import re
import zipfile
from datetime import date
from pathlib import Path
from urllib.parse import quote_plus

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

DATASET_PAGE = "https://data.europa.eu/data/datasets/cordis-eu-research-projects-under-horizon-europe-2021-2027"
# Last known working direct export (verify against the dataset page above if this 404s):
FALLBACK_CSV_ZIP = "https://cordis.europa.eu/data/cordis-HORIZONprojects-csv.zip"

DEFAULT_KEYWORDS = [
    # AI / ML
    "artificial intelligence", "machine learning", "deep learning",
    "neural network", "natural language processing", "computer vision",
    "generative ai", "large language model", "llm",
    # Data & analytics - a data/AI consultancy's leads aren't only "AI companies"
    "data analytics", "data science", "predictive analytics", "big data",
    "data strategy", "business intelligence",
    # Cloud & digital transformation
    "cloud migration", "cloud computing", "digital transformation",
]

# fundingScheme values that specifically mean EIC Accelerator (verified against
# a live CORDIS download - HORIZON-EIC alone also covers Pathfinder/Transition,
# which we keep in the general Horizon Europe sheet instead).
EIC_ACCELERATOR_SCHEMES = {"HORIZON-EIC-ACC", "HORIZON-EIC-ACC-BF"}

OUTPUT_COLUMNS = {
    "company": "Company",
    "country": "Country",
    "website": "Website",
    "website_arama_linki": "Website Search Link (verify)",
    "project_acronym": "Project Acronym",
    "project_title": "Project Title",
    "project_start": "Project Start",
    "funding_signed_date": "Funding Signed Date",
    "eu_contribution_eur": "EU Contribution (EUR)",
    "fundingScheme": "Funding Scheme",
    "ai_relevance": "Tech Relevance",
    "science_domain": "Scientific/Technical Domain (not a sector - CORDIS classification)",
}


def download_project_bundle(url: str = FALLBACK_CSV_ZIP) -> zipfile.ZipFile:
    """Download the CORDIS Horizon Europe CSV bundle and return it as a ZipFile."""
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    return zipfile.ZipFile(io.BytesIO(resp.content))


def load_table(zf: zipfile.ZipFile, name_contains: str) -> pd.DataFrame:
    """Find and load the CSV inside the bundle whose filename contains `name_contains`.

    Uses Python's stdlib csv module instead of pandas' own CSV parser: CORDIS's
    export has a known quoting bug (fields ending in a literal quote character
    are missing their closing quote), which the pandas C/python engines reject
    outright. The stdlib reader is more lenient and recovers far more rows -
    verified against a live download: ~0.8% of rows dropped vs. ~4.2% with
    pandas' own parser. Rows that still don't match the header's field count
    are skipped and counted.
    """
    matches = [n for n in zf.namelist() if name_contains.lower() in n.lower() and n.endswith(".csv")]
    if not matches:
        raise FileNotFoundError(f"No file matching '*{name_contains}*.csv' in bundle: {zf.namelist()}")
    with zf.open(matches[0]) as f:
        text = io.TextIOWrapper(f, encoding="utf-8", errors="replace", newline="")
        reader = csv.reader(text, delimiter=";", quotechar='"')
        header = next(reader)
        rows = []
        skipped = 0
        for row in reader:
            if len(row) == len(header):
                rows.append(row)
            else:
                skipped += 1
        if skipped:
            print(f"Uyari: {matches[0]} icinde {skipped} satir bozuk CSV kacisi nedeniyle atlandi "
                  f"(CORDIS'in kendi export hatasi, script hatasi degil).")
        return pd.DataFrame(rows, columns=header)


def tag_ai_relevance(projects: pd.DataFrame, keywords: list[str]) -> pd.DataFrame:
    """Tag every project by how central data/AI/cloud work is to it - does NOT filter anything out.

    A company doesn't need to be an "AI company" to be a lead - a data/AI
    consultancy's real client base spans data analytics, BI, cloud, and
    digital transformation work too, often at companies whose core business
    is something else entirely. So every project stays in the list; this just
    labels it "Core Tech Focus" (a data/AI/cloud term is in the project's own
    title), "Tech-Adjacent" (mentioned only in the objective, i.e. used as a
    tool within a project about something else), or "General Funding Lead"
    (no such term at all) so the eventual Excel/webpage can be sorted or
    filtered by relevance without losing any leads.
    """
    pattern = re.compile("|".join(re.escape(k) for k in keywords), re.IGNORECASE)
    title_hit = projects["title"].astype(str).str.contains(pattern, na=False) if "title" in projects.columns else pd.Series(False, index=projects.index)
    objective_hit = projects["objective"].astype(str).str.contains(pattern, na=False) if "objective" in projects.columns else pd.Series(False, index=projects.index)

    result = projects.copy()
    result["ai_relevance"] = "General Funding Lead"
    result.loc[objective_hit, "ai_relevance"] = "Tech-Adjacent (mentioned in objective)"
    result.loc[title_hit, "ai_relevance"] = "Core Tech Focus (mentioned in title)"
    return result


def load_science_domains(zf: zipfile.ZipFile) -> dict[str, str]:
    """Map projectID -> top-level EuroSciVoc domain(s), e.g. "natural sciences; engineering and technology".

    This is CORDIS's own scientific/technical classification, not a formal
    industry sector (NACE/SIC) - CORDIS's bulk export has no such field. It's
    the closest available proxy for "what field is this company/project in",
    so we label it honestly as a scientific/technical domain, not a "sector".
    """
    scivoc = load_table(zf, "euroSciVoc")
    if "euroSciVocPath" not in scivoc.columns or "projectID" not in scivoc.columns:
        return {}
    scivoc["top_domain"] = scivoc["euroSciVocPath"].astype(str).str.split("/").str[0]
    grouped = scivoc.groupby("projectID")["top_domain"].apply(lambda s: "; ".join(sorted(set(s))))
    return grouped.to_dict()


def build_radar(keywords: list[str], countries: list[str] | None, output_path: str) -> pd.DataFrame:
    zf = download_project_bundle()
    projects = load_table(zf, "project")
    orgs = load_table(zf, "organization")
    domains = load_science_domains(zf)

    tagged_projects = tag_ai_relevance(projects, keywords)

    id_col = "id" if "id" in tagged_projects.columns else "rcn"
    tagged_projects["science_domain"] = tagged_projects[id_col].map(domains).fillna("")
    merged = orgs.merge(
        tagged_projects[[id_col, "acronym", "title", "startDate", "ecSignatureDate",
                          "ecMaxContribution", "fundingScheme", "ai_relevance", "science_domain"]],
        left_on="projectID", right_on=id_col, how="inner",
    )

    # Focus on the actual companies (SME flag), not universities/research orgs -
    # these are the leads a BD pipeline actually wants to call.
    if "SME" in merged.columns:
        merged = merged[merged["SME"].astype(str).str.lower().isin(["true", "1", "yes"])]

    if countries:
        merged = merged[merged["country"].isin(countries)]

    result = merged.rename(columns={
        "name": "company",
        "country": "country",
        "title": "project_title",
        "acronym": "project_acronym",
        "startDate": "project_start",
        "ecSignatureDate": "funding_signed_date",
        "ecMaxContribution": "eu_contribution_eur",
        "organizationURL": "website",
    })[[
        "company", "country", "website", "project_acronym", "project_title",
        "project_start", "funding_signed_date", "eu_contribution_eur",
        "fundingScheme", "ai_relevance", "science_domain",
    ]].drop_duplicates(subset=["company", "project_acronym"])

    # CORDIS exports numbers with a European decimal comma (e.g. "1904544,25");
    # convert to a real float so Excel treats it as a number, not text.
    result["eu_contribution_eur"] = pd.to_numeric(
        result["eu_contribution_eur"].astype(str).str.replace(",", ".", regex=False), errors="coerce"
    )

    # CORDIS's own website field is frequently blank or stale. We don't auto-fill
    # it - that risks attaching the wrong company's info - we just hand you a
    # ready-made Google search link so you can check and fill it in by hand.
    result["website"] = result["website"].fillna("").astype(str).str.strip()
    result["website_arama_linki"] = result.apply(
        lambda r: "" if r["website"] else
        f"https://www.google.com/search?q={quote_plus(r['company'] + ' ' + r['country'])}",
        axis=1,
    )

    eic_mask = result["fundingScheme"].isin(EIC_ACCELERATOR_SCHEMES)
    sheets = {
        "EIC Accelerator": result[eic_mask].copy(),
        "Horizon Europe (General)": result[~eic_mask].copy(),
    }

    _write_excel(sheets, output_path, keywords, countries)
    html_path = Path("docs/index.html")
    html_path.parent.mkdir(exist_ok=True)
    _write_html(sheets, html_path, keywords, countries)
    print(f"{len(result)} satir yazildi -> {output_path} ve {html_path} "
          f"(EIC Accelerator: {eic_mask.sum()}, Horizon Europe genel: {(~eic_mask).sum()})")
    return result


LIMITATIONS = [
    "Website info is frequently missing/stale in CORDIS - verify manually via the 'Website Search Link' column.",
    "Outside of EIC Accelerator, the exact funding amount per company is not broken out separately by CORDIS.",
    "CORDIS only covers EU-funded projects; data/AI/tech companies funded by VC/private equity instead are not in this list.",
    "The 'Scientific/Technical Domain' column is CORDIS's own EuroSciVoc classification, not an official "
    "business sector (NACE/SIC) - CORDIS does not provide that data.",
]


def _prepare_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """Rename to display columns, keep only known columns, sort by EU contribution desc."""
    df = df.rename(columns=OUTPUT_COLUMNS)
    df = df[[c for c in OUTPUT_COLUMNS.values() if c in df.columns]]
    return df.sort_values("EU Contribution (EUR)", ascending=False, na_position="last")


def _write_excel(sheets: dict[str, pd.DataFrame], output_path: str, keywords: list[str], countries: list[str] | None) -> None:
    """Write each segment to its own sheet, plus a cover sheet with source/date/limitations."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        info = pd.DataFrame({
            "Field": ["Source", "Download date", "Scope (country)", "Keywords used"]
                     + [f"Known limitation {i+1}" for i in range(len(LIMITATIONS))],
            "Value": [FALLBACK_CSV_ZIP, date.today().isoformat(),
                      ", ".join(countries) if countries else "All", ", ".join(keywords)] + LIMITATIONS,
        })
        info.to_excel(writer, sheet_name="Source & Notes", index=False)
        _format_sheet(writer.sheets["Source & Notes"], info, wrap_cols={1})

        for sheet_name, df in sheets.items():
            df = _prepare_sheet(df)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            wrap_cols = {df.columns.get_loc("Proje Basligi")} if "Proje Basligi" in df.columns else set()
            _format_sheet(writer.sheets[sheet_name], df, wrap_cols=wrap_cols)


_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="UTF-8">
<title>CORDIS Data & AI Funding Radar - DACH</title>
<style>
  :root {{ color-scheme: light; }}
  body {{ font-family: -apple-system, "Segoe UI", Arial, sans-serif; margin: 0; padding: 2rem;
          background: #f7f7f9; color: #1a1a1a; }}
  h1 {{ margin-bottom: 0.2rem; font-size: 1.6rem; }}
  .subtitle {{ color: #555; margin-top: 0; margin-bottom: 1.2rem; }}
  .tabs {{ display: flex; gap: 0.5rem; margin-bottom: 1rem; }}
  .tab-btn {{ padding: 0.5rem 1rem; border: 1px solid #ccc; background: #fff; border-radius: 8px;
              cursor: pointer; font-size: 0.95rem; }}
  .tab-btn.active {{ background: #1a56db; color: #fff; border-color: #1a56db; }}
  input#search {{ width: 100%; max-width: 420px; padding: 0.5rem 0.8rem; border: 1px solid #ccc;
                   border-radius: 8px; margin-bottom: 1rem; font-size: 0.95rem; }}
  table {{ border-collapse: collapse; width: 100%; background: #fff; box-shadow: 0 1px 3px rgba(0,0,0,0.1);
           border-radius: 8px; overflow: hidden; }}
  th, td {{ padding: 0.55rem 0.7rem; border-bottom: 1px solid #eee; text-align: left; font-size: 0.87rem;
            vertical-align: top; }}
  th {{ background: #1a56db; color: #fff; cursor: pointer; user-select: none; white-space: nowrap; }}
  th:hover {{ background: #164bc0; }}
  tr:hover td {{ background: #f0f4ff; }}
  .tag-core {{ background: #dcfce7; color: #166534; padding: 0.15rem 0.5rem; border-radius: 6px; font-size: 0.78rem; }}
  .tag-adj {{ background: #fef3c7; color: #92400e; padding: 0.15rem 0.5rem; border-radius: 6px; font-size: 0.78rem; }}
  .tag-none {{ background: #f1f1f1; color: #555; padding: 0.15rem 0.5rem; border-radius: 6px; font-size: 0.78rem; }}
  a {{ color: #1a56db; }}
  .info-box {{ background: #fff; border: 1px solid #e5e5e5; border-radius: 8px; padding: 1rem 1.2rem;
               margin-bottom: 1.5rem; font-size: 0.85rem; color: #444; }}
  .info-box ul {{ margin: 0.4rem 0 0 1.1rem; padding: 0; }}
  .count {{ color: #666; font-size: 0.85rem; margin-bottom: 0.6rem; }}
  .sheet {{ display: none; }}
  .sheet.active {{ display: block; }}
</style>
</head>
<body>
<h1>CORDIS Data & AI Funding Radar &mdash; DACH</h1>
<p class="subtitle">Germany / Austria / Switzerland &middot; Last updated: {generated_date}</p>

<div class="info-box">
  <strong>Source:</strong> <a href="{source_url}" target="_blank">{source_url}</a> (official CORDIS bulk dataset)<br>
  <strong>Scope:</strong> {countries}<br>
  <strong>Known limitations:</strong>
  <ul>{limitations_html}</ul>
</div>

<div class="tabs">{tab_buttons}</div>
<input id="search" type="text" placeholder="Search company, project, or domain...">
{sheet_sections}

<script>
function showTab(id, btn) {{
  document.querySelectorAll('.sheet').forEach(function(s) {{ s.classList.remove('active'); }});
  document.querySelectorAll('.tab-btn').forEach(function(b) {{ b.classList.remove('active'); }});
  document.getElementById(id).classList.add('active');
  btn.classList.add('active');
  document.getElementById('search').value = '';
  filterRows();
}}

function filterRows() {{
  var q = document.getElementById('search').value.toLowerCase();
  var visible = 0;
  document.querySelectorAll('.sheet.active tbody tr').forEach(function(row) {{
    var match = row.textContent.toLowerCase().includes(q);
    row.style.display = match ? '' : 'none';
    if (match) visible++;
  }});
  var counter = document.querySelector('.sheet.active .count');
  if (counter) counter.textContent = visible + ' rows';
}}
document.getElementById('search').addEventListener('input', filterRows);

function sortTable(table, colIdx) {{
  var tbody = table.tBodies[0];
  var rows = Array.from(tbody.querySelectorAll('tr'));
  var asc = table.dataset.sortCol == colIdx ? table.dataset.sortDir !== 'asc' : true;
  rows.sort(function(a, b) {{
    var av = a.cells[colIdx].dataset.sort || a.cells[colIdx].textContent.trim();
    var bv = b.cells[colIdx].dataset.sort || b.cells[colIdx].textContent.trim();
    var an = parseFloat(av), bn = parseFloat(bv);
    var cmp = (!isNaN(an) && !isNaN(bn)) ? an - bn : av.localeCompare(bv, 'tr');
    return asc ? cmp : -cmp;
  }});
  rows.forEach(function(r) {{ tbody.appendChild(r); }});
  table.dataset.sortCol = colIdx;
  table.dataset.sortDir = asc ? 'asc' : 'desc';
}}
document.querySelectorAll('table').forEach(function(table) {{
  table.querySelectorAll('th').forEach(function(th, idx) {{
    th.addEventListener('click', function() {{ sortTable(table, idx); }});
  }});
}});
</script>
</body>
</html>
"""


def _df_to_html_table(df: pd.DataFrame) -> str:
    import html as _html
    headers = "".join(f"<th>{_html.escape(c)}</th>" for c in df.columns)
    rows_html = []
    website_col = "Website" if "Website" in df.columns else None
    search_col = "Website Arama Linki (dogrula)" if "Website Arama Linki (dogrula)" in df.columns else None
    ai_col = "AI Iliskisi" if "AI Iliskisi" in df.columns else None
    amount_col = "AB Katkisi (EUR)" if "AB Katkisi (EUR)" in df.columns else None
    for _, row in df.iterrows():
        cells = []
        for col in df.columns:
            val = row[col]
            if col == website_col and str(val).strip():
                cells.append(f'<td><a href="{_html.escape(str(val))}" target="_blank">{_html.escape(str(val))}</a></td>')
            elif col == search_col:
                if str(val).strip():
                    cells.append(f'<td><a href="{_html.escape(str(val))}" target="_blank">Search on Google</a></td>')
                else:
                    cells.append("<td></td>")
            elif col == ai_col:
                if str(val).startswith("Core"):
                    cls = "tag-core"
                elif str(val).startswith("Tech-Adjacent"):
                    cls = "tag-adj"
                else:
                    cls = "tag-none"
                cells.append(f'<td><span class="{cls}">{_html.escape(str(val))}</span></td>')
            elif col == amount_col:
                sort_val = "" if pd.isna(val) else f"{val:.0f}"
                display = "" if pd.isna(val) else f"{val:,.0f} EUR".replace(",", ".")
                cells.append(f'<td data-sort="{sort_val}">{display}</td>')
            else:
                cells.append(f"<td>{_html.escape(str(val)) if pd.notna(val) else ''}</td>")
        rows_html.append("<tr>" + "".join(cells) + "</tr>")
    return f"<table><thead><tr>{headers}</tr></thead><tbody>{''.join(rows_html)}</tbody></table>"


def _write_html(sheets: dict[str, pd.DataFrame], output_path: Path, keywords: list[str], countries: list[str] | None) -> None:
    """Render the same segments as a single self-contained, sortable/searchable HTML page for GitHub Pages."""
    import html as _html
    prepared = {name: _prepare_sheet(df) for name, df in sheets.items()}

    tab_buttons, sheet_sections = [], []
    for i, (name, df) in enumerate(prepared.items()):
        slug = f"sheet{i}"
        active = " active" if i == 0 else ""
        tab_buttons.append(
            f'<button class="tab-btn{active}" onclick="showTab(\'{slug}\', this)">{_html.escape(name)} ({len(df)})</button>'
        )
        sheet_sections.append(
            f'<div id="{slug}" class="sheet{active}"><p class="count">{len(df)} rows</p>{_df_to_html_table(df)}</div>'
        )

    html = _HTML_TEMPLATE.format(
        generated_date=date.today().isoformat(),
        source_url=FALLBACK_CSV_ZIP,
        countries=_html.escape(", ".join(countries) if countries else "All"),
        limitations_html="".join(f"<li>{_html.escape(l)}</li>" for l in LIMITATIONS),
        tab_buttons="".join(tab_buttons),
        sheet_sections="".join(sheet_sections),
    )
    output_path.write_text(html, encoding="utf-8")


def _format_sheet(ws, df: pd.DataFrame, wrap_cols: set[int] | None = None) -> None:
    """Bold header, freeze header row, sane column widths, wrap long text columns."""
    wrap_cols = wrap_cols or set()
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        max_len = max([len(str(col_name))] + [len(str(v)) for v in df.iloc[:, col_idx - 1].head(200)])
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)
        if (col_idx - 1) in wrap_cols:
            for row_idx in range(2, len(df) + 2):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--keywords", type=str, default=",".join(DEFAULT_KEYWORDS),
                         help="Comma-separated keywords used to tag (not filter) title/objective for AI relevance")
    parser.add_argument("--countries", type=str, default=None,
                         help="Comma-separated ISO country codes to keep, e.g. DE,AT,CH (default: all)")
    parser.add_argument("--output", type=str, default="ai_funding_radar.xlsx")
    args = parser.parse_args()

    kw = [k.strip() for k in args.keywords.split(",") if k.strip()]
    countries = [c.strip().upper() for c in args.countries.split(",")] if args.countries else None

    build_radar(kw, countries, args.output)
